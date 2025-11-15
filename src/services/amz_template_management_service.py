"""
Amazon Template Management Service
亚马逊类目模板管理服务

负责编排更新亚马逊类目模板的整个业务流程，包括基于报错的规则矫正
"""
from src.services.amz_template_parser import AdvancedTemplateParser
from src.repositories.amz_template_repository import AmzTemplateRepository
from sqlalchemy.orm import Session
import os
import logging
from typing import Tuple, Dict, List, Any, Set
import openpyxl
import re

logger = logging.getLogger(__name__)

# ========================================================================
# 常量定义
# ========================================================================

# 内部属性名到亚马逊字段名的映射
INTERNAL_TO_AMZ_MAP = {
    "color_name": ["Color", "Color Name", "Main Color"],
    "size_name": ["Size", "Size Name", "Apparel Size", "Ring Size", "Shoe Size"],
    "material_name": ["Material", "Main Material", "Material Type"],
    "style_name": ["Style", "Style Name"],
    "item_package_quantity": ["Item Package Quantity", "Number Of Items"]
}

# 默认的高优先级变体主题
DEFAULT_PRIORITY_THEMES = [
    "COLOR/SIZE", 
    "COLOR", 
    "SIZE", 
    "MATERIAL", 
    "STYLE", 
    "COLOR/STYLE"
]


class TemplateManagementService:
    """
    服务层
    
    负责编排更新亚马逊类目模板的整个业务流程，包括基于报错的规则矫正
    """

    def __init__(self, db: Session):
        """
        初始化服务
        
        Args:
            db: SQLAlchemy Session 对象
        """
        self.db = db
        self.repo = AmzTemplateRepository(db=self.db)

    # ========================================================================
    # 3.2 功能：更新模板
    # ========================================================================

    def update_template_from_file(
        self, 
        file_path: str, 
        category: str
    ) -> Tuple[bool, str]:
        """
        处理单个模板文件的核心业务逻辑
        
        流程：
        1. 解析 Excel 模板文件
        2. 生成变体属性映射
        3. 确定优先级主题
        4. 保存到数据库
        
        Args:
            file_path: Excel 模板文件的完整路径
            category: 品类名称（如 'HOME_MIRROR', 'CABINET'）
            
        Returns:
            元组 (操作是否成功, 相关消息)
        """
        logger.info(
            f"🚀 开始处理模板文件 '{file_path}'，品类为 '{category}'..."
        )

        # 验证文件存在
        if not file_path or not os.path.exists(file_path):
            message = f"文件不存在或路径无效: {file_path}"
            logger.error(message)
            return False, message

        # 步骤1：解析模板文件
        logger.info("调用解析器模块...")
        parser = AdvancedTemplateParser(file_path)
        parse_success, parse_message = parser.parse()

        if not parse_success:
            logger.error(f"模板解析失败: {parse_message}")
            return False, f"模板解析失败: {parse_message}"

        results = parser.get_results()

        # 步骤2：生成变体属性映射
        logger.info("开始自动生成变体属性映射...")
        template_fields = results.get("fields", [])
        variation_themes = parser.get_all_variation_themes()
        variation_mapping = self._generate_variation_mapping(
            template_fields, 
            variation_themes
        )
        results["variation_mapping"] = variation_mapping

        # 步骤3：确定优先级主题
        priority_themes = self._determine_priority_themes(category)
        results["priority_themes"] = priority_themes

        # 步骤4：保存到数据库
        logger.info("调用仓库层以保存数据...")
        template_name = os.path.basename(file_path)

        inserted_id = self.repo.save_parsed_data(
            category=category,
            template_name=template_name,
            results=results
        )

        if inserted_id is not None:
            message = (
                f"模板 '{template_name}' 已成功处理并存入数据库 "
                f"(ID: {inserted_id})。"
            )
            logger.info(message)
            print(
                f"\n✅ 成功! 最终为品类 '{category}' "
                f"保存的高优先级主题为: {priority_themes}"
            )
            return True, message
        else:
            message = "数据保存到数据库失败，请检查日志获取详细信息。"
            logger.error(message)
            return False, message

    # ========================================================================
    # 3.3 功能：矫正规则
    # ========================================================================

    def correct_rules_from_report(
        self, 
        report_path: str, 
        category: str
    ) -> Tuple[bool, str]:
        """
        从亚马逊报错文件中自动矫正模板的必填字段规则
        
        流程：
        1. 解析报错文件，提取需要修正的字段名
        2. 从数据库获取当前的字段定义
        3. 矫正字段定义并写回数据库
        
        Args:
            report_path: 亚马逊报错文件 (.xlsm) 的完整路径
            category: 该文件对应的品类名称
            
        Returns:
            元组 (操作是否成功, 相关消息)
        """
        print(f"\n🚀 启动模板规则自动矫正流程，品类: '{category}'...")

        # 步骤 1: 解析报错文件，提取需要修正的字段名
        try:
            print(
                f"➡️ 步骤 1/3: 正在解析报错文件 "
                f"'{os.path.basename(report_path)}'..."
            )
            required_fields = self._parse_report_for_required_fields(report_path)
            
            if not required_fields:
                message = (
                    "✅ 解析完成，但未在报错文件中找到与'必填项缺失'"
                    "相关的错误 (Error code 90220)。无需矫正。"
                )
                print(message)
                return True, message
                
            print(f"✔️ 从报告中识别出 {len(required_fields)} 个必填字段。")

        except Exception as e:
            message = f"❌ 解析报错文件时失败: {e}"
            logger.exception(message)
            print(message)
            return False, message

        # 步骤 2: 从数据库获取当前的字段定义
        print("➡️ 步骤 2/3: 正在从数据库获取当前模板规则...")
        db_result = self.repo.find_latest_template_id_and_defs(category)
        
        if not db_result:
            message = (
                f"❌ 未能在数据库中找到品类 '{category}' 的模板记录，"
                "无法执行矫正。"
            )
            print(message)
            return False, message

        record_id, field_definitions = db_result
        print("✔️ 成功获取数据库记录。")

        # 步骤 3: 矫正字段定义并写回数据库
        try:
            print("➡️ 步骤 3/3: 正在矫正规则并更新数据库...")
            updated_defs, corrected_fields = self._apply_corrections(
                field_definitions, 
                required_fields
            )

            if not corrected_fields:
                message = (
                    "✅ 所有在报错文件中提及的必填字段，"
                    "在数据库中已是必填状态。无需矫正。"
                )
                print(message)
                return True, message

            success = self.repo.update_field_definitions_by_id(
                record_id, 
                updated_defs
            )
            
            if success:
                self.db.commit()
                final_message = (
                    f"✅ 成功！已为品类 '{category}' 矫正了 "
                    f"{len(corrected_fields)} 个字段的必填规则:\n   - "
                    + "\n   - ".join(sorted(list(corrected_fields)))
                )
                print(final_message)
                return True, final_message
            else:
                self.db.rollback()
                message = "❌ 更新数据库时发生错误，操作已回滚。"
                print(message)
                return False, message
                
        except Exception as e:
            self.db.rollback()
            message = f"❌ 在矫正和更新过程中发生未知错误: {e}"
            logger.exception(message)
            print(message)
            return False, message

    # ========================================================================
    # 内部辅助方法
    # ========================================================================

    def _generate_variation_mapping(
        self, 
        template_fields: List[str], 
        variation_themes: List[str]
    ) -> Dict[str, str]:
        """
        生成变体属性映射
        
        将内部属性名（如 color_name）映射到模板中的实际字段名
        
        Args:
            template_fields: 模板中的所有字段列表
            variation_themes: 变体主题列表
            
        Returns:
            变体映射字典，例如 {'color_name': 'Color', 'size_name': 'Size'}
        """
        variation_mapping = {}
        
        # 从变体主题中识别可能的变体属性字段
        possible_variation_fields = set()
        for theme in variation_themes:
            parts = theme.split('/')
            possible_variation_fields.update(p.strip().lower() for p in parts)
            
        logger.info(
            f"从变体主题中识别出可能的变体属性字段 (小写): "
            f"{possible_variation_fields}"
        )
        
        # 创建模板字段的小写映射
        template_fields_lower = {
            field.lower(): field 
            for field in template_fields
        }
        
        # 匹配内部属性名到模板字段
        for internal_key, amz_name_variations in INTERNAL_TO_AMZ_MAP.items():
            for amz_name in amz_name_variations:
                amz_name_lower = amz_name.lower()
                
                # 必须同时满足：在模板中存在 且 是变体属性
                if (amz_name_lower in template_fields_lower and 
                    amz_name_lower in possible_variation_fields):
                    
                    original_cased_field = template_fields_lower[amz_name_lower]
                    variation_mapping[internal_key] = original_cased_field
                    logger.info(
                        f"成功映射: 内部键 '{internal_key}' -> "
                        f"模板列 '{original_cased_field}'"
                    )
                    break
                    
        logger.info(
            f"为该模板生成的最终 variation_mapping: {variation_mapping}"
        )
        return variation_mapping

    def _determine_priority_themes(self, category: str) -> List[str]:
        """
        确定优先级主题
        
        优先级：用户输入 > 历史配置 > 系统默认
        
        Args:
            category: 品类名称
            
        Returns:
            高优先级主题列表
        """
        print("\n--- 变体主题优先级配置 ---")
        print("💡 请输入此品类的高优先级变体主题, 用逗号','分隔。")
        print("   例如: COLOR/SIZE, COLOR, STYLE")
        print("   直接按 Enter 键可跳过，系统将尝试自动沿用旧配置或使用默认值。")
        user_input = input("请输入: ").strip()

        # 用户输入
        if user_input:
            themes = [
                theme.strip().upper() 
                for theme in user_input.split(',') 
                if theme.strip()
            ]
            print(f"✅ 已采纳您的输入: {themes}")
            return themes

        # 历史配置
        print("ℹ️ 您已跳过输入。正在检查历史配置...")
        latest_themes = self.repo.find_latest_priority_themes_by_category(category)
        
        if latest_themes:
            print(f"✅ 已成功沿用上个版本的配置: {latest_themes}")
            return latest_themes

        # 系统默认
        print("⚠️ 未找到历史配置。将使用系统默认的高优先级列表。")
        print(f"   默认列表为: {DEFAULT_PRIORITY_THEMES}")
        return DEFAULT_PRIORITY_THEMES

    def _parse_report_for_required_fields(self, file_path: str) -> Set[str]:
        """
        解析亚马逊报错 .xlsm 文件，提取因 'is required but not supplied' 
        导致的错误字段名
        
        Args:
            file_path: 报错文件路径
            
        Returns:
            必填字段名的集合
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"指定的报错文件路径不存在: {file_path}")

        required_fields = set()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if "Feed Processing Summary" not in wb.sheetnames:
            raise ValueError(
                "文件中未找到名为 'Feed Processing Summary' 的工作表。"
            )

        sheet = wb["Feed Processing Summary"]

        # 动态查找表头
        header = []
        header_row_idx = -1
        
        for i in range(1, 10):  # 扫描前10行
            row_values = [
                str(cell.value).strip() 
                for cell in sheet[i] 
                if cell.value
            ]
            if "Error code" in row_values and "Error message" in row_values:
                header = [
                    str(cell.value).strip() if cell.value else "" 
                    for cell in sheet[i]
                ]
                header_row_idx = i
                break

        if header_row_idx == -1:
            raise ValueError(
                "未能在 'Feed Processing Summary' 中找到包含 "
                "'Error code' 和 'Error message' 的表头。"
            )

        # 创建列名到索引的映射
        try:
            code_col_idx = header.index("Error code")
            msg_col_idx = header.index("Error message")
        except ValueError:
            raise ValueError(
                "表头中必须同时包含 'Error code' 和 'Error message' 列。"
            )

        # 使用正则表达式从错误消息中提取字段名
        pattern = re.compile(r"'(.+?)' is required but not supplied\.")

        for row in sheet.iter_rows(min_row=header_row_idx + 1):
            error_code = row[code_col_idx].value
            error_message = row[msg_col_idx].value

            if (str(error_code).strip() == '90220' and 
                isinstance(error_message, str)):
                match = pattern.search(error_message)
                if match:
                    field_name = match.group(1)
                    required_fields.add(field_name)

        return required_fields

    def _apply_corrections(
        self, 
        definitions: Dict, 
        fields_to_correct: Set[str]
    ) -> Tuple[Dict, Set[str]]:
        """
        将从报错文件中解析出的字段名应用到从数据库获取的字段定义中
        
        Args:
            definitions: 字段定义字典
            fields_to_correct: 需要矫正的字段名集合
            
        Returns:
            元组 (更新后的定义字典, 实际矫正的字段集合)
        """
        corrected_fields = set()
        
        # 创建一个从 local_label 到其父级键的映射
        label_to_key_map = {
            v.get("local_label"): k
            for k, v in definitions.items()
            if v and isinstance(v, dict) and v.get("local_label")
        }

        for field_name in fields_to_correct:
            target_key = label_to_key_map.get(field_name)
            
            if target_key:
                field_def = definitions.get(target_key, {})
                
                # 检查是否真的需要更新
                if (field_def.get("required_child") != "Required" or 
                    field_def.get("required_single") != "Required"):
                    
                    field_def["required_child"] = "Required"
                    field_def["required_single"] = "Required"
                    corrected_fields.add(field_name)
            else:
                logger.warning(
                    f"在数据库的字段定义中，未能找到 local_label 为 "
                    f"'{field_name}' 的记录，已跳过。"
                )

        return definitions, corrected_fields