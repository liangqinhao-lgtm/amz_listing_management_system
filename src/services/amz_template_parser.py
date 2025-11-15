"""
Amazon Template Parser
亚马逊类目模板解析器

负责解析 Amazon 品类模板 Excel 文件，提取字段、定义和有效值
"""
import openpyxl
import os
import logging
from collections import defaultdict
import traceback
from typing import List, Dict, Any, Tuple

logger = logging.getLogger(__name__)


class AdvancedTemplateParser:
    """
    高级模板解析器
    
    负责解析 Amazon 品类模板 Excel 文件，提取：
    - 字段列表（fields）
    - 字段定义（field_definitions）
    - 有效值（valid_values）
    - 变体主题（variation_themes）
    """

    def __init__(self, file_path: str, skip_deprecated: bool = True):
        """
        构造函数
        
        Args:
            file_path: 要解析的模板文件路径
            skip_deprecated: 是否跳过标记为废弃的值
        """
        self.file_path = file_path
        self.skip_deprecated = skip_deprecated
        self.fields: List[str] = []
        self.field_definitions: Dict[str, Any] = {}
        self.valid_values: List[Dict[str, Any]] = []
        self.wb = None
        logger.info(f"解析器已为文件 '{os.path.basename(file_path)}' 初始化.")

    def _log_and_print(self, message: str, level: str = "info"):
        """统一的日志记录方法"""
        if level == "info":
            logger.info(message)
        elif level == "warning":
            logger.warning(message)
        elif level == "error":
            logger.error(message)
        else:
            logger.debug(message)

    def _open_workbook(self) -> bool:
        """打开 Excel 工作簿"""
        try:
            self._log_and_print(f"📂 正在打开工作簿: {self.file_path}")
            self.wb = openpyxl.load_workbook(
                self.file_path, 
                read_only=True, 
                data_only=True
            )
            self._log_and_print(f"📋 工作表列表: {', '.join(self.wb.sheetnames)}")
            return True
        except Exception as e:
            self._log_and_print(f"❌ 打开工作簿失败: {e}", "error")
            return False

    def parse(self) -> Tuple[bool, str]:
        """
        主解析流程
        
        Returns:
            元组 (是否成功, 消息)
        """
        try:
            if not self._open_workbook():
                return False, "无法打开Excel工作簿"

            if not self._parse_template_sheet():
                return False, "解析 'Template' 工作表失败"

            if not self._parse_data_definitions_advanced():
                return False, "解析 'Data Definitions' 工作表失败"

            if not self._parse_valid_values():
                return False, "解析 'Valid Values' 工作表失败"

            if not self.fields or not self.field_definitions:
                return False, "解析结果不完整：未能提取到字段或字段定义"

            summary = (
                f"字段数量: {len(self.fields)}, "
                f"字段定义数量: {len(self.field_definitions)}, "
                f"有效属性数量: {len(self.valid_values)}"
            )
            self._log_and_print(f"✅ 解析成功。摘要: {summary}")
            return True, "解析成功"
            
        except Exception as e:
            error_trace = traceback.format_exc()
            self._log_and_print(
                f"❌ 解析过程中发生异常: {e}\n{error_trace}", 
                "error"
            )
            return False, f"解析失败: {e}"
        finally:
            if self.wb:
                self.wb.close()

    def get_results(self) -> Dict[str, Any]:
        """
        获取结构化的解析结果
        
        Returns:
            包含 fields, field_definitions, valid_values 的字典
        """
        return {
            "fields": self.fields,
            "field_definitions": self.field_definitions,
            "valid_values": self.valid_values
        }

    def get_all_variation_themes(self) -> List[str]:
        """
        从解析结果中，专门提取出所有可用的变体主题 (Variation Themes)
        
        这是生成 variation_mapping 的关键步骤之一
        
        Returns:
            包含所有变体主题字符串的列表，例如 ["Color", "Size", "Color-Size"]
        """
        for valid_value_group in self.valid_values:
            # 亚马逊模板中，定义变体主题的属性名通常是 'Variation Theme Name'
            if valid_value_group.get("attribute") == "Variation Theme Name":
                themes = valid_value_group.get("values", [])
                logger.info(f"在 'Valid Values' 中找到 {len(themes)} 个可用变体主题。")
                return themes
                
        logger.warning(
            "在 'Valid Values' 中未找到 'Variation Theme Name' 属性，"
            "无法提取变体主题。"
        )
        return []

    def _parse_template_sheet(self) -> bool:
        """解析 Template 表获取字段列表"""
        self._log_and_print("🔍 开始解析 'Template' 工作表...")
        sheet_name = "Template"
        
        if sheet_name not in self.wb.sheetnames:
            self._log_and_print(
                f"❌ 工作簿中不存在 '{sheet_name}' 工作表", 
                "error"
            )
            return False

        sheet = self.wb[sheet_name]
        field_rows_to_try = [4, 3, 2, 1]

        for row_idx in field_rows_to_try:
            if row_idx > sheet.max_row:
                continue

            row = [cell.value for cell in sheet[row_idx]]
            if not row:
                continue

            fields = [
                str(val).strip() 
                for val in row 
                if val is not None and str(val).strip()
            ]
            
            if fields:
                self.fields = fields
                self._log_and_print(
                    f"✅ 在第 {row_idx} 行找到 {len(fields)} 个字段。"
                )
                return True

        self._log_and_print(
            "⚠️ 未在 'Template' 表中找到有效的字段行。", 
            "warning"
        )
        return False

    def _parse_data_definitions_advanced(self) -> bool:
        """
        解析 Data Definitions 表，动态查找表头行
        
        Returns:
            是否解析成功
        """
        self._log_and_print("🔍 开始解析 'Data Definitions' 工作表...")
        sheet_name = "Data Definitions"
        
        if sheet_name not in self.wb.sheetnames:
            self._log_and_print(
                f"❌ 工作簿中不存在 '{sheet_name}' 工作表", 
                "error"
            )
            return False

        sheet = self.wb[sheet_name]

        # 动态表头查找逻辑
        header_row_idx = -1
        raw_headers = []
        # 定义必须存在的关键列，用于识别表头行
        required_headers = {"Field Name", "Local Label Name"}

        # 扫描前5行来查找表头
        for i in range(1, 6):
            if i > sheet.max_row:
                break

            current_row_values = {
                str(cell.value).strip() 
                for cell in sheet[i] 
                if cell.value
            }
            
            # 如果当前行包含了所有必需的列名，就认定为表头行
            if required_headers.issubset(current_row_values):
                header_row_idx = i
                raw_headers = [
                    str(cell.value).strip() if cell.value else "" 
                    for cell in sheet[i]
                ]
                self._log_and_print(
                    f"✅ 在第 {header_row_idx} 行找到 'Data Definitions' 的表头。"
                )
                break

        if header_row_idx == -1:
            self._log_and_print(
                f"❌ 未能找到有效的表头行。扫描了前 {min(5, sheet.max_row)} 行。", 
                "error"
            )
            return False

        # 列名映射
        header_variations = {
            "group": ["Group Name"],
            "field_name": ["Field Name"],
            "local_label": ["Local Label Name", "Local Label"],
            "accepted_values": ["Accepted Values"],
            "example": ["Example"],
            "required_parent": ["Required for Parent?", "Required for Parant?"],
            "required_child": ["Required for Child?"],
            "required_single": [
                "Required for single SKU product?", 
                "Required for single SKU"
            ]
        }

        column_mapping = {}
        for key, variations in header_variations.items():
            for variation in variations:
                try:
                    idx = raw_headers.index(variation)
                    column_mapping[key] = idx
                    break
                except ValueError:
                    continue

        if 'field_name' not in column_mapping:
            self._log_and_print(
                f"❌ 表头行中未找到必需的 'Field Name' 列。"
                f"检测到的表头: {raw_headers}", 
                "error"
            )
            return False

        self._log_and_print(
            f"✅ 成功映射 'Data Definitions' 的列: {list(column_mapping.keys())}"
        )

        current_group = ""
        # 从表头行的下一行开始解析数据
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            if not any(v is not None for v in row_values):
                continue

            group_name_idx = column_mapping.get("group", -1)
            group_name = (
                str(row_values[group_name_idx]).strip() 
                if group_name_idx != -1 and row_values[group_name_idx] 
                else ""
            )

            field_name_idx = column_mapping["field_name"]
            field_name = (
                str(row_values[field_name_idx]).strip() 
                if row_values[field_name_idx] 
                else ""
            )

            # 处理分组行
            if group_name and not field_name:
                current_group = group_name
                continue

            # 处理字段行
            if field_name and field_name.lower() != "field name":
                field_def = {
                    "group": current_group, 
                    "field_name": field_name
                }
                
                for key, idx in column_mapping.items():
                    if key not in field_def and idx < len(row_values):
                        field_def[key] = (
                            str(row_values[idx]) 
                            if row_values[idx] is not None 
                            else ""
                        )

                self.field_definitions[field_name] = field_def
                continue

        self._log_and_print(
            f"✅ 'Data Definitions' 解析完成，"
            f"找到 {len(self.field_definitions)} 个字段定义。"
        )
        return True

    def _parse_valid_values(self) -> bool:
        """解析 Valid Values 表"""
        self._log_and_print("🔍 开始解析 'Valid Values' 工作表...")
        sheet_name = "Valid Values"
        
        if sheet_name not in self.wb.sheetnames:
            self._log_and_print(
                f"ℹ️ 工作簿中不存在 '{sheet_name}' 工作表，跳过。", 
                "info"
            )
            return True

        sheet = self.wb[sheet_name]
        current_group = None

        for row_idx in range(1, sheet.max_row + 1):
            row = [
                str(cell.value).strip() if cell.value else "" 
                for cell in sheet[row_idx]
            ]
            if not any(row):
                continue

            # 处理分组行
            if row[0]:
                current_group = row[0]
                continue

            # 处理属性行
            if row[1] and "[" in row[1] and "]" in row[1]:
                attr_declaration = row[1]
                try:
                    attr_name_part, scope_part = attr_declaration.rsplit("[", 1)
                    scope = scope_part.split("]", 1)[0].strip()
                    attr_name = attr_name_part.strip().rstrip("-").strip()
                except ValueError:
                    attr_name = attr_declaration
                    scope = "UNKNOWN"

                valid_vals = [
                    val 
                    for val in row[2:] 
                    if val and not self._is_deprecated(val)
                ]

                if valid_vals:
                    self.valid_values.append({
                        "group": current_group,
                        "attribute": attr_name,
                        "scope": scope,
                        "values": valid_vals
                    })

        self._log_and_print(
            f"✅ 'Valid Values' 解析完成，"
            f"找到 {len(self.valid_values)} 个有效属性。"
        )
        return True

    def _is_deprecated(self, value: str) -> bool:
        """
        检查值是否被标记为废弃
        
        Args:
            value: 要检查的值
            
        Returns:
            是否为废弃值
        """
        if not self.skip_deprecated:
            return False
            
        val_lower = value.lower()
        deprecated_terms = ["deprecated", "do not use", "obsolete"]
        return any(term in val_lower for term in deprecated_terms)