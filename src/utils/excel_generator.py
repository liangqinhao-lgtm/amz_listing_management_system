"""
Excel Generator
生成Amazon发品Excel文件
"""
import logging
import openpyxl
import datetime
import uuid
from pathlib import Path
from typing import Dict, Any, List
from collections import defaultdict

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Excel发品文件生成器
    
    职责：
    - 读取品类Excel模板（.xlsm）
    - 填充产品数据到模板
    - 生成最终的发品文件
    """
    
    def __init__(self, template_base_path: Path = None, output_base_path: Path = None):
        """
        初始化生成器
        
        Args:
            template_base_path: 模板文件基础目录，默认自动查找
            output_base_path: 输出文件基础目录，默认为 output/
        """
        if template_base_path is None:
            template_base_path = self._find_template_path()
        
        if output_base_path is None:
            output_base_path = self._find_output_path()
        
        self.template_base_path = template_base_path
        self.output_base_path = output_base_path
        
        # 确保输出目录存在
        self.output_base_path.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Excel生成器初始化完成")
        logger.info(f"  模板目录: {self.template_base_path}")
        logger.info(f"  输出目录: {self.output_base_path}")
    
    def _find_template_path(self) -> Path:
        """查找模板文件目录"""
        # 从当前文件向上查找项目根目录
        current = Path(__file__).resolve()
        
        for parent in current.parents:
            template_dir = parent / "template_files"
            if template_dir.exists():
                return template_dir
        
        raise FileNotFoundError("未找到 template_files 目录")
    
    def _find_output_path(self) -> Path:
        """查找输出目录"""
        current = Path(__file__).resolve()
        
        for parent in current.parents:
            if (parent / "template_files").exists():
                return parent / "output"
        
        # 默认在项目根创建output
        return Path("output")
    
    def generate_excel(
        self,
        rows_data: List[Dict[str, Any]],
        category_name: str,
        batch_id: uuid.UUID
    ) -> str:
        """
        生成Excel发品文件
        
        Args:
            rows_data: 数据行列表，每个元素是字段名到值的映射
            category_name: 品类名称（如 'CABINET', 'HOME_MIRROR'）
            batch_id: 批次ID
            
        Returns:
            生成的文件完整路径
            
        Raises:
            FileNotFoundError: 模板文件不存在
            Exception: 生成过程出错
        """
        if not rows_data:
            raise ValueError("rows_data 不能为空")
        
        logger.info(f"开始生成 {category_name} 的Excel文件，共 {len(rows_data)} 行数据")
        
        # 1. 加载模板
        template_filename = f"{category_name.upper()}.xlsm"
        template_path = self.template_base_path / template_filename
        
        if not template_path.exists():
            raise FileNotFoundError(f"找不到品类 '{category_name}' 的模板文件: {template_path}")
        
        logger.info(f"加载模板: {template_path}")
        
        try:
            # 使用keep_vba保留宏
            wb = openpyxl.load_workbook(template_path, keep_vba=True)
            ws = wb["Template"]
            
            # 2. 解析表头
            header_row_index = 4
            data_start_row_index = 7
            
            header_map = self._parse_header(ws, header_row_index)
            
            logger.info(f"解析到 {len(header_map)} 个表头字段")
            
            # 3. 填充数据
            self._fill_data(ws, rows_data, header_map, data_start_row_index)
            
            # 4. 生成输出文件名
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            batch_short = str(batch_id)[:8]
            output_filename = f"AmazonUpload_{category_name}_{timestamp}_batch_{batch_short}.xlsm"
            output_path = self.output_base_path / output_filename
            
            # 5. 保存文件
            wb.save(output_path)
            
            logger.info(f"✅ 成功生成Excel文件: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"❌ 生成Excel文件失败: {e}", exc_info=True)
            raise
    
    def _parse_header(self, worksheet, header_row_index: int) -> Dict[str, List[int]]:
        """
        解析表头，建立字段名到列索引的映射
        
        处理重复列名的情况（如多个 Bullet Point 列）
        
        Args:
            worksheet: Excel工作表
            header_row_index: 表头所在行号（1-based）
            
        Returns:
            字典，键为字段名，值为列索引列表
        """
        header_map = defaultdict(list)
        
        # 读取表头行
        for col_idx in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=header_row_index, column=col_idx).value
            
            if header_value:
                header_map[header_value].append(col_idx)
        
        return dict(header_map)
    
    def _fill_data(
        self,
        worksheet,
        rows_data: List[Dict[str, Any]],
        header_map: Dict[str, List[int]],
        data_start_row: int
    ):
        """
        填充数据到工作表
        
        Args:
            worksheet: Excel工作表
            rows_data: 数据行列表
            header_map: 表头映射
            data_start_row: 数据起始行号
        """
        # 需要提醒的硬编码字段集合
        special_fields = {
            'Parentage Level',
            'Child Relationship Type',
            'Parent SKU',
            'Variation Theme Name'
        }

        missing_special_fields = set()

        for row_idx, row_data in enumerate(rows_data):
            current_row = data_start_row + row_idx
            
            for field_name, value in row_data.items():
                col_indices = header_map.get(field_name)
                
                if not col_indices:
                    # 字段在模板中不存在，跳过
                    logger.debug(f"字段 '{field_name}' 在模板中不存在，跳过")
                    # 如果是硬编码字段，记录用于汇总提醒
                    if field_name in special_fields:
                        missing_special_fields.add(field_name)
                    continue
                
                # 处理列表类型的值（如 Bullet Point）
                if isinstance(value, list):
                    for item_idx, item_value in enumerate(value):
                        if item_idx < len(col_indices):
                            col_idx = col_indices[item_idx]
                            worksheet.cell(row=current_row, column=col_idx, value=item_value)
                else:
                    # 单个值，写入第一个匹配的列
                    col_idx = col_indices[0]
                    worksheet.cell(row=current_row, column=col_idx, value=value)
        
        logger.info(f"填充完成，共写入 {len(rows_data)} 行数据")

        # 汇总提醒：模板缺少应写入的硬编码字段
        if missing_special_fields:
            for f in sorted(missing_special_fields):
                logger.warning(f"应写入字段 {f}，模板文档不包含该字段，请检查原因。")
